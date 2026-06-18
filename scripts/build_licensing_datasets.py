#!/usr/bin/env python3
"""
Build clean, reproducible licensing datasets from the SAS PRR C264029 release.

Inputs (raw PRR files, committed under data/prr-responses/C264029/):
  - 2014-2025PDR.xlsx            individual dog-license issuances (items #1, #2)
  - SAS_Licensing_Data_2014-25_revenue.xlsx   annual dog-license revenue (item #4)

Outputs (data/):
  - licensing-by-year.csv     dog licenses issued per year + new/renewal + altered share
  - licensing-by-zip.csv      dog licenses by home ZIP (pooled) + neighborhood label
  - licensing-revenue.csv     annual dog-license revenue (partial years flagged)

Caveats baked into the data, not just the docs:
  - Each row is a license ISSUED on a date, not a snapshot of active licenses.
  - 2025 new-vs-renewal is unreliable (SAS moved to 1-year-only; renewals reissued
    as "new"). Total 2025 count is fine.
  - ZIP is the owner's home address.

Usage: .venv/bin/python3 scripts/build_licensing_datasets.py
"""
from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
import pandas as pd

REPO = Path(__file__).resolve().parent.parent
SRC = REPO / "data/prr-responses/C264029"
PDR = SRC / "2014-2025PDR.xlsx"
REV = SRC / "SAS_Licensing_Data_2014-25_revenue.xlsx"
OUT = REPO / "data"

# Home-ZIP -> neighborhood label (Seattle). Approximate, for display only.
ZIP_NEIGH = {
    "98115": "NE Seattle / Wedgwood", "98103": "Wallingford / Fremont",
    "98117": "Ballard", "98125": "Lake City / Northgate", "98118": "Columbia City",
    "98116": "West Seattle / Alki", "98199": "Magnolia", "98122": "Capitol Hill / CD",
    "98126": "West Seattle / Delridge", "98107": "Ballard / Fremont",
    "98144": "Beacon Hill / Mt Baker", "98106": "South Delridge / Highland Park",
    "98105": "University District", "98112": "Madison Park / Montlake",
    "98136": "West Seattle / Fauntleroy", "98109": "Queen Anne / SLU",
    "98119": "Queen Anne", "98102": "Capitol Hill / Eastlake",
    "98108": "Georgetown / Beacon Hill", "98133": "Bitter Lake / Broadview",
    "98178": "Rainier Beach", "98146": "White Center / Highland Park",
    "98121": "Belltown", "98101": "Downtown", "98104": "Pioneer Square / ID",
    "98177": "Broadview / Blue Ridge", "98168": "Tukwila-adjacent",
}


def build_by_year(dogs: pd.DataFrame) -> None:
    g = dogs.groupby("year")
    rows = []
    for y, sub in g:
        n = len(sub)
        renew = (sub["Renewal Yes/No"] == "Yes").sum()
        new = (sub["Renewal Yes/No"] == "No").sum()
        altered = (sub["Altered"] == "Yes").sum()
        rows.append({
            "year": int(y),
            "licenses_issued": n,
            "new_licenses": int(new),
            "renewals": int(renew),
            "pct_altered": round(altered / n * 100, 1),
            "note": "new/renewal unreliable: 1-yr transition" if int(y) == 2025 else "",
        })
    prov = ("calculated: scripts/build_licensing_datasets.py from PRR C264029 "
            "(2014-2025PDR.xlsx); each row = a license issued, not active count")
    with open(OUT / "licensing-by-year.csv", "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()) + ["provenance"])
        w.writeheader()
        for r in rows:
            r["provenance"] = prov
            w.writerow(r)
    print(f"Wrote licensing-by-year.csv ({len(rows)} years)")


def build_by_zip(dogs: pd.DataFrame) -> None:
    z = dogs["Postal Code"].astype(str).str.extract(r"(\d{5})")[0]
    counts = z.value_counts()
    prov = ("calculated: scripts/build_licensing_datasets.py from PRR C264029; "
            "home ZIP, 2014-2025 pooled issuances (not per-capita)")
    with open(OUT / "licensing-by-zip.csv", "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["zip", "neighborhood", "licenses_2014_25", "provenance"])
        for zc, c in counts.items():
            if not zc or zc == "nan":
                continue
            w.writerow([zc, ZIP_NEIGH.get(zc, ""), int(c), prov])
    print(f"Wrote licensing-by-zip.csv ({(counts.index != 'nan').sum()} ZIPs)")


def build_revenue() -> None:
    wb = openpyxl.load_workbook(REV, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    # row order: [title], [year header], [Dog Licenses + values], [Total], notes...
    dog_row = next(r for r in rows if r and str(r[0]).strip() == "Dog Licenses")
    years = [2014, 2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025]
    vals = list(dog_row[1:13])
    partial = {2015: "actuals through Nov + projected Dec", 2017: "incomplete"}
    prov = ("sourced: PRR C264029 SAS_Licensing_Data_2014-25_revenue.xlsx; "
            "dog-license General Fund revenue; funds the whole Animal Shelter, not OLAs")
    with open(OUT / "licensing-revenue.csv", "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["year", "dog_license_revenue", "partial_flag", "provenance"])
        for y, v in zip(years, vals):
            if v is None:
                continue
            w.writerow([y, round(float(v)), partial.get(y, ""), prov])
    print(f"Wrote licensing-revenue.csv")


def main() -> None:
    df = pd.read_excel(PDR, sheet_name=0)
    df["year"] = pd.to_datetime(df["Date Issued"], errors="coerce").dt.year
    dogs = df[(df["Species"] == "Dog") & df["year"].notna()].copy()
    print(f"Loaded {len(dogs):,} dog-license rows, {int(dogs['year'].min())}-{int(dogs['year'].max())}")
    build_by_year(dogs)
    build_by_zip(dogs)
    build_revenue()


if __name__ == "__main__":
    main()
