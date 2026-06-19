#!/usr/bin/env python3
"""
End-to-end verification harness for the enforcement data pipeline.

Run this after every change to:
  - any XLSX in data/prr-responses/C049204/ or data/prr-responses/C263949/
  - scripts/build_enforcement_datasets.py
  - data/enforcement-citations.csv (do not hand-edit; this script will fail)
  - any of the small derived CSVs:
      data/enforcement-hotspots.csv
      data/enforcement-hotspots-extra.csv
      data/enforcement-by-park-year.csv
      data/enforcement-year-metrics.csv

It re-derives every published number from the raw PRR workbooks and the
consolidated CSV and asserts equality. Any drift fails the script with a
specific message identifying the broken invariant.

Usage:
  cd <repo root> && .venv/bin/python scripts/verify_enforcement_data.py

Exit code 0 = all assertions passed; non-zero = at least one failure.
"""

from __future__ import annotations

import csv
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
PRR_OLD_DIR = REPO_ROOT / "data" / "prr-responses" / "C049204"
PRR_NEW_DIR = REPO_ROOT / "data" / "prr-responses" / "C263949"
CITATIONS = REPO_ROOT / "data" / "enforcement-citations.csv"
BY_PARK_YEAR = REPO_ROOT / "data" / "enforcement-by-park-year.csv"
HOTSPOTS = REPO_ROOT / "data" / "enforcement-hotspots.csv"
HOTSPOTS_EXTRA = REPO_ROOT / "data" / "enforcement-hotspots-extra.csv"

# Ground-truth values pinned from the source PRR responses.
# Anchor every published number in the repo to one of these constants.

# C049204 — DLP-only Jan 2014 → Oct 15 2019.
# Pre-2019 counts derived directly from the original XLSX sheet header
# row counts (header included). Post-build, the 2019 rows from this PRR
# are dropped by build_enforcement_datasets.py — see C263949 row.
GROUND_TRUTH_OLD_PRR_DLP_BY_YEAR = {
    "2014": 183,
    "2015": 519,
    "2016": 952,
    "2017": 844,
    "2018": 1276,
    # 2019 rows from this PRR are intentionally NOT in the consolidated
    # CSV — superseded by C263949 full-year 2019. The 2019 partial-year
    # row count in the original file was 1029, used as a delta check
    # against C263949's 2019 DLP total.
}
OLD_PRR_2019_PARTIAL_DLP_COUNT = 1029

# C263949 — all parks-related violations Jan 2019 → Apr 17 2026.
# Per-file row counts pinned to the "Total Violations: N" sentinel that
# SAS embedded in column 48 of every data row.
GROUND_TRUTH_NEW_PRR_PER_FILE = {
    "CaseViolationDetail-2019-2020_Release.xlsx": 1806,
    "CaseViolationDetail-2021-2022_Release.xlsx": 758,
    "CaseViolationDetail-2023-2024_Release.xlsx": 799,
    "CaseViolationDetail-2025-2026YTD_Release.xlsx": 395,
}
# Per-year totals across the new PRR (all categories, not DLP-only).
# Re-derive from raw XLSX before each release if the underlying files
# are re-issued by SAS; these are the numbers we assert against.
GROUND_TRUTH_NEW_PRR_ALL_BY_YEAR = {
    "2019": 1360,
    "2020": 446,
    "2021": 559,
    "2022": 199,
    "2023": 285,
    "2024": 514,
    "2025": 326,
    "2026": 69,
}
GROUND_TRUTH_NEW_PRR_DLP_BY_YEAR = {
    "2019": 1181,
    "2020": 393,
    "2021": 471,
    "2022": 169,
    "2023": 248,
    "2024": 447,
    "2025": 267,
    "2026": 65,
}

# Fee tiers per SMC 18.12.080. Verified against multiple citation rows.
FEE_TIERS_DLP = {1: 54, 2: 109, 3: 136, 4: 162}
FEE_TIER_LICENSE = 125

failures: list[str] = []


def check(condition: bool, message: str) -> None:
    if condition:
        print(f"  PASS  {message}")
    else:
        failures.append(message)
        print(f"  FAIL  {message}")


# ---- Raw-file sanity checks -----------------------------------------------

def verify_new_prr_raw() -> None:
    """Re-read each C263949 XLSX, count data rows, compare to sentinel + ground truth."""
    print("\n[1] C263949 — raw XLSX row counts and sentinels")
    if not PRR_NEW_DIR.exists():
        check(False, f"directory does not exist: {PRR_NEW_DIR}")
        return
    for path in sorted(PRR_NEW_DIR.glob("*.xlsx")):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        sentinel: int | None = None
        count = 0
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 49:
                continue
            if not isinstance(row[2], (int, float)):
                continue
            count += 1
            cell = row[48]
            if isinstance(cell, str) and cell.startswith("Total Violations:"):
                try:
                    sentinel = int(cell.split(":")[1].strip())
                except (IndexError, ValueError):
                    pass
        wb.close()
        gt = GROUND_TRUTH_NEW_PRR_PER_FILE.get(path.name)
        check(sentinel == count, f"{path.name}: row count {count} == sentinel {sentinel}")
        check(gt is not None and count == gt, f"{path.name}: row count {count} == ground truth {gt}")


def verify_old_prr_raw() -> None:
    """Re-read C049204 XLSX, count DLP rows by year, compare to ground truth."""
    print("\n[2] C049204 — raw DLP row counts per year")
    if not PRR_OLD_DIR.exists():
        check(False, f"directory does not exist: {PRR_OLD_DIR}")
        return
    year_counts: Counter = Counter()
    for path in sorted(PRR_OLD_DIR.glob("*.xlsx")):
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers: list[str] | None = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c).strip() if c is not None else "" for c in row]
                    continue
                if all(c is None for c in row):
                    continue
                d = dict(zip(headers, row))
                dt = d.get("Issue Date/Time")
                if isinstance(dt, datetime):
                    year_counts[str(dt.year)] += 1
                elif dt:
                    s = str(dt)
                    if s[:4].isdigit():
                        year_counts[s[:4]] += 1
        wb.close()

    for year, expected in GROUND_TRUTH_OLD_PRR_DLP_BY_YEAR.items():
        actual = year_counts.get(year, 0)
        check(actual == expected, f"C049204 {year}: {actual} == {expected}")
    actual_2019 = year_counts.get("2019", 0)
    check(
        actual_2019 == OLD_PRR_2019_PARTIAL_DLP_COUNT,
        f"C049204 2019 (Jan-Oct 15, partial): {actual_2019} == {OLD_PRR_2019_PARTIAL_DLP_COUNT}",
    )


# ---- Consolidated CSV checks ----------------------------------------------

def load_citations() -> list[dict]:
    if not CITATIONS.exists():
        return []
    with CITATIONS.open() as fh:
        return list(csv.DictReader(fh))


def verify_citations_csv(rows: list[dict]) -> None:
    print("\n[3] Consolidated citations CSV — internal consistency")
    if not rows:
        check(False, "data/enforcement-citations.csv is empty or missing")
        return

    # Expected total rows: old DLP (2014-2018) + new (full).
    expected_total = sum(GROUND_TRUTH_OLD_PRR_DLP_BY_YEAR.values()) + sum(
        GROUND_TRUTH_NEW_PRR_PER_FILE.values()
    )
    check(len(rows) == expected_total, f"row count: {len(rows)} == {expected_total}")

    # Source-PRR split.
    by_source = Counter(r["source_prr"] for r in rows)
    check(
        by_source["C049204"] == sum(GROUND_TRUTH_OLD_PRR_DLP_BY_YEAR.values()),
        f"C049204 rows kept: {by_source['C049204']} == {sum(GROUND_TRUTH_OLD_PRR_DLP_BY_YEAR.values())}",
    )
    check(
        by_source["C263949"] == sum(GROUND_TRUTH_NEW_PRR_PER_FILE.values()),
        f"C263949 rows ingested: {by_source['C263949']} == {sum(GROUND_TRUTH_NEW_PRR_PER_FILE.values())}",
    )

    # No 2019 rows from C049204 (the overlap rule).
    leftover_old_2019 = sum(1 for r in rows if r["source_prr"] == "C049204" and r["year"] == "2019")
    check(leftover_old_2019 == 0, f"C049204 2019 rows present: 0 expected, found {leftover_old_2019}")

    # By-year totals (all categories).
    yc_all = Counter(r["year"] for r in rows)
    for year, expected in GROUND_TRUTH_OLD_PRR_DLP_BY_YEAR.items():
        check(yc_all[year] == expected, f"all-categories year {year}: {yc_all[year]} == {expected}")
    for year, expected in GROUND_TRUTH_NEW_PRR_ALL_BY_YEAR.items():
        check(yc_all[year] == expected, f"all-categories year {year}: {yc_all[year]} == {expected}")

    # By-year totals (DLP only).
    yc_dlp = Counter(r["year"] for r in rows if r["dlp_only"] == "True")
    for year, expected in GROUND_TRUTH_OLD_PRR_DLP_BY_YEAR.items():
        check(yc_dlp[year] == expected, f"DLP-only year {year}: {yc_dlp[year]} == {expected}")
    for year, expected in GROUND_TRUTH_NEW_PRR_DLP_BY_YEAR.items():
        check(yc_dlp[year] == expected, f"DLP-only year {year}: {yc_dlp[year]} == {expected}")

    # Schema sanity.
    required = {
        "year",
        "offense_level",
        "violation_item",
        "violation_category",
        "dlp_only",
        "source_prr",
        "fee",
        "case_result",
        "location_canon",
    }
    missing = required - set(rows[0].keys())
    check(not missing, f"required columns present: missing={sorted(missing)}")


def verify_fee_arithmetic(rows: list[dict]) -> None:
    """Each fee should match the SMC fee tier for its offense level (DLP only)."""
    print("\n[4] Fee arithmetic — per-tier consistency")
    fee_by_level: dict[int, Counter] = defaultdict(Counter)
    for r in rows:
        if r["dlp_only"] != "True":
            continue
        if r["case_result"] not in ("Citation",):
            continue
        try:
            ol = int(r["offense_level"])
        except (ValueError, TypeError):
            continue
        try:
            fee = int(float(r["fee"])) if r["fee"] not in ("", None) else 0
        except (ValueError, TypeError):
            continue
        fee_by_level[ol][fee] += 1

    for level, expected_fee in FEE_TIERS_DLP.items():
        # Most paid rows at a level should sit at the SMC tier.
        if level not in fee_by_level:
            continue
        total = sum(fee_by_level[level].values())
        at_tier = fee_by_level[level][expected_fee]
        if total == 0:
            continue
        pct = at_tier / total
        check(
            pct >= 0.85,
            f"DLP offense level {level}: {pct:.0%} of paid rows are at ${expected_fee} ({at_tier}/{total})",
        )

    # Compute total assessed fee revenue (sum of fee column over Citation rows).
    total_revenue = 0
    for r in rows:
        try:
            fee = float(r["fee"]) if r["fee"] not in ("", None) else 0.0
        except (ValueError, TypeError):
            continue
        total_revenue += fee
    print(f"  INFO  total assessed fee revenue across consolidated CSV: ${int(total_revenue):,}")


# ---- Derived CSV checks ---------------------------------------------------

def _dlp_park_named_counts(rows: list[dict]) -> Counter:
    """DLP-only park-named citation counts per canonical park, full 2014–2026."""
    return Counter(
        r["location_canon"]
        for r in rows
        if r["dlp_only"] == "True"
        and r["location_type"] == "park_named"
        and r["location_canon"]
    )


def verify_hotspots(rows: list[dict]) -> None:
    print("\n[5] enforcement-hotspots.csv reconciles to consolidated DLP park-named counts (2014–2026)")
    if not HOTSPOTS.exists():
        check(False, "enforcement-hotspots.csv missing")
        return
    park_counts = _dlp_park_named_counts(rows)
    ranked = [p for p, _ in park_counts.most_common()]
    with HOTSPOTS.open() as fh:
        csv_rows = list(csv.DictReader(fh))
    # Each CSV count must equal the recomputed DLP park-named count exactly
    # (the CSV is now generated from the same data, no tolerance needed).
    bad = [(r["park"], int(r["count"]), park_counts.get(r["park"], 0))
           for r in csv_rows if int(r["count"]) != park_counts.get(r["park"], 0)]
    check(not bad, f"hotspots counts match consolidated DLP park-named: mismatches={bad}")
    # Every CSV park must be in the genuine top-20 by count.
    top20 = set(ranked[:20])
    not_top = [r["park"] for r in csv_rows if r["park"] not in top20]
    check(not not_top, f"all hotspots parks are in the top-20: outliers={not_top}")
    counts_seq = [int(r["count"]) for r in csv_rows]
    check(counts_seq == sorted(counts_seq, reverse=True), "hotspots.csv sorted by count desc")


def verify_hotspots_extra(rows: list[dict]) -> None:
    print("\n[6] enforcement-hotspots-extra.csv reconciles to consolidated DLP park-named counts (ranks 21–40)")
    if not HOTSPOTS_EXTRA.exists():
        check(False, "enforcement-hotspots-extra.csv missing")
        return
    park_counts = _dlp_park_named_counts(rows)
    ranked = [p for p, _ in park_counts.most_common()]
    with HOTSPOTS_EXTRA.open() as fh:
        csv_rows = list(csv.DictReader(fh))
    bad = [(r["park"], int(r["count"]), park_counts.get(r["park"], 0))
           for r in csv_rows if int(r["count"]) != park_counts.get(r["park"], 0)]
    check(not bad, f"hotspots-extra counts match consolidated DLP park-named: mismatches={bad}")
    band = set(ranked[20:40])
    not_band = [r["park"] for r in csv_rows if r["park"] not in band]
    check(not not_band, f"all hotspots-extra parks are in ranks 21–40: outliers={not_band}")


def verify_by_park_year(rows: list[dict]) -> None:
    print("\n[7] enforcement-by-park-year.csv reconciles to consolidated citations CSV")
    if not BY_PARK_YEAR.exists():
        check(False, "enforcement-by-park-year.csv missing")
        return
    counts = Counter(
        (r["location_canon"], r["year"], r["dlp_only"]) for r in rows if r["location_canon"] and r["year"]
    )
    sample_failed = 0
    with BY_PARK_YEAR.open() as fh:
        derived = list(csv.DictReader(fh))
    for d in derived:
        key = (d["park"], d["year"], d["dlp_only"])
        expected = counts.get(key, 0)
        actual = int(d["citations"])
        if actual != expected:
            sample_failed += 1
            if sample_failed <= 5:
                print(f"  FAIL  by-park-year {key}: {actual} != {expected}")
    check(sample_failed == 0, f"by-park-year all rows match citations CSV ({len(derived)} rows checked)")

    # Total rows in the by-park-year aggregate should equal the consolidated
    # CSV's per-(park,year,dlp_only) count of non-empty rows.
    expected_keys = sum(1 for _ in counts)
    check(
        len(derived) == expected_keys,
        f"by-park-year row count {len(derived)} == unique (park,year,dlp_only) triples {expected_keys}",
    )


YEAR_METRICS = REPO_ROOT / "data" / "enforcement-year-metrics.csv"


def verify_year_metrics(rows: list[dict]) -> None:
    """Recompute the per-year metrics CSV from the consolidated CSV + the
    staffing/cost model in build_enforcement_metrics.py and assert equality."""
    print("\n[8] enforcement-year-metrics.csv reconciles to citations CSV + cost model")
    if not YEAR_METRICS.exists():
        check(False, "enforcement-year-metrics.csv missing")
        return

    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "build_enforcement_metrics", REPO_ROOT / "scripts" / "build_enforcement_metrics.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(mod)
    STAFFING = mod.STAFFING
    FAS = mod.FAS_ACO_ANNUAL
    FMW = mod.FMW_ANNUAL
    PARTIAL = mod.PARTIAL_YEARS

    dlp_by_year = Counter()
    off_by_year: dict[str, Counter] = defaultdict(Counter)
    rev_by_year: dict[str, float] = defaultdict(float)
    for r in rows:
        y = r["year"]
        if not y:
            continue
        if r["dlp_only"] == "True":
            dlp_by_year[y] += 1
            if r["offense_level"]:
                try:
                    off_by_year[y][int(r["offense_level"])] += 1
                except ValueError:
                    pass
        if r["fee"] not in ("", None):
            try:
                rev_by_year[y] += float(r["fee"])
            except (TypeError, ValueError):
                pass

    with YEAR_METRICS.open() as fh:
        committed = {row["year"]: row for row in csv.DictReader(fh)}

    for y, (aco, fmw) in STAFFING.items():
        m = committed.get(y)
        if not m:
            check(False, f"year-metrics missing row for {y}")
            continue
        exp_cost = round(aco * FAS + fmw * FMW)
        exp_dlp = dlp_by_year.get(y, 0)
        check(int(m["dlp_citations"]) == exp_dlp, f"year-metrics {y} dlp: {m['dlp_citations']} == {exp_dlp}")
        check(int(m["annual_cost"]) == exp_cost, f"year-metrics {y} cost: {m['annual_cost']} == {exp_cost}")
        check(round(rev_by_year.get(y, 0.0)) == int(m["fee_revenue"]), f"year-metrics {y} revenue: {m['fee_revenue']} == {round(rev_by_year.get(y,0.0))}")
        if y not in PARTIAL and exp_dlp:
            exp_cpc = round(exp_cost / exp_dlp)
            check(int(m["cost_per_citation"]) == exp_cpc, f"year-metrics {y} cost/citation: {m['cost_per_citation']} == {exp_cpc}")
            exp_fte = round(exp_dlp / (aco + fmw), 1)
            check(abs(float(m["citations_per_fte"]) - exp_fte) < 0.05, f"year-metrics {y} per-FTE: {m['citations_per_fte']} == {exp_fte}")
        else:
            check(m["cost_per_citation"] == "", f"year-metrics {y} cost/citation blank (partial year)")
        off_total = sum(off_by_year[y].values())
        if off_total:
            exp_first = round(100 * off_by_year[y].get(1, 0) / off_total, 1)
            check(abs(float(m["first_offense_pct"]) - exp_first) < 0.05, f"year-metrics {y} first-offense%: {m['first_offense_pct']} == {exp_first}")

    # Cumulative cost-recovery cross-check (the 11% / $3.34M headline)
    cum_cost = sum(round(STAFFING[y][0]*FAS + STAFFING[y][1]*FMW) for y in STAFFING)
    cum_rev = round(sum(rev_by_year.values()))
    recovery = 100 * cum_rev / cum_cost
    check(3_300_000 <= cum_cost <= 3_400_000, f"cumulative cost ~$3.34M: ${cum_cost:,}")
    check(cum_rev == 351099, f"cumulative revenue == $351,099: ${cum_rev:,}")
    check(10.0 <= recovery <= 11.0, f"cost recovery 10-11%: {recovery:.1f}%")
    print(f"  INFO  cumulative cost ${cum_cost:,}, revenue ${cum_rev:,}, recovery {recovery:.1f}%")


# ---- HTML prose check -----------------------------------------------------

# The CSV checks above guarantee the *data* is correct. They do NOT guarantee
# that the hardcoded numbers written into the enforcement page's prose match
# that data — every figure typed into a <p> or stat card is a separate place a
# stale or fat-fingered number can hide (this is how "$1,716" survived once
# when the real value was $1,730, and how a "33%" top-10 share survived when
# the data said 40%). This check recomputes each load-bearing prose figure from
# the consolidated CSV + metrics CSV + OLA acreage and asserts the formatted
# string actually appears in the rendered HTML. It is deliberately string-based
# ("grep the page for the number the data implies") rather than DOM-parsing, so
# it stays simple and dependency-free.
#
# The corpus is defined by an explicit ALLOWLIST, not a glob, for two reasons:
#  1. "Genuine public site" should mean exactly the canonical, linked, indexed
#     pages — an allowlist says so unambiguously, and it FAILS loudly if one of
#     those pages goes missing or is renamed (a glob would just silently shrink
#     the corpus and let an error slip through).
#  2. New content is published in two phases. While a change is staged (e.g. the
#     working enforcement-draft.html we're about to promote), its numbers must
#     already be guarded — so staged files are scanned too, kept in a separate
#     STAGED list. Once a staged file is promoted into a public page (draft →
#     enforcement.html), delete it from STAGED: the scan then verifies the
#     updated complete public site, with the content living in its public page.
DOCS_DIR = REPO_ROOT / "docs"
OLAS_CSV = REPO_ROOT / "data" / "seattle-olas.csv"

# The genuine public site: canonical, linked, indexed pages (see CLAUDE.md).
PUBLIC_PAGES = [
    "index.html",
    "part1-the-gap.html",
    "part2-access.html",
    "part3.html",
    "enforcement.html",
    "budget.html",
    "peer-cities.html",
    "opinion.html",
    "updates.html",
]
# Content staged for the next push but not yet promoted into a public page.
# Scanned now so the numbers we're about to publish are already guarded.
# ADD an entry while a page is staged (e.g. "enforcement-draft.html"); REMOVE it
# the moment it is promoted into a public page above, so the scan tracks the real
# public site. Currently empty: the enforcement rebuild is live in enforcement.html.
STAGED_PAGES: list[str] = []


def _site_corpus() -> tuple[str, list[str], list[str]]:
    """Return (combined_text, scanned_filenames, missing_public_pages)."""
    parts: list[str] = []
    scanned: list[str] = []
    missing: list[str] = []
    for name in PUBLIC_PAGES:
        p = DOCS_DIR / name
        if p.exists():
            parts.append(p.read_text())
            scanned.append(name)
        else:
            missing.append(name)
    for name in STAGED_PAGES:
        p = DOCS_DIR / name
        if p.exists():
            parts.append(p.read_text())
            scanned.append(f"{name} (staged)")
    return "\n".join(parts), scanned, missing


def verify_html_prose(rows: list[dict]) -> None:
    print("\n[9] site HTML prose — hardcoded numbers match the data (public site + staged)")
    html, pages, missing = _site_corpus()
    for name in missing:
        check(False, f"public page present in docs/: {name}")
    if not html:
        print("  INFO  no public/staged HTML pages found — skipping prose check")
        return
    print(f"  INFO  scanning {len(pages)} pages: {', '.join(pages)}")

    # ---- recompute every prose figure from the data ----
    dlp = [r for r in rows if r["dlp_only"] == "True"]
    named = [r for r in dlp if r.get("location_type") == "park_named" and r["location_canon"]]

    def yr(r: dict) -> int:
        return int(r["year"]) if r["year"] else 0

    dlp_total = len(dlp)
    dlp_by_year = Counter(r["year"] for r in dlp)
    all_by_year = Counter(r["year"] for r in rows if r["year"])
    peak_2018 = dlp_by_year["2018"]
    # "Best year since the COVID period" = strongest year AFTER the shaded
    # 2020-2021 COVID window (2021's 471 partial-rebound is inside it).
    best_post = max((dlp_by_year[str(y)] for y in range(2022, 2027)), default=0)
    pct_2024_of_peak = round(100 * best_post / peak_2018)
    drop_2020 = round(100 * (dlp_by_year["2020"] - dlp_by_year["2019"]) / dlp_by_year["2019"])
    non_dlp_post2019 = sum(1 for r in rows if r["dlp_only"] != "True" and yr(r) >= 2019)

    # metrics CSV
    with YEAR_METRICS.open() as fh:
        ym = {row["year"]: row for row in csv.DictReader(fh)}
    cpc_2018 = int(ym["2018"]["cost_per_citation"])
    cpc_2022 = int(ym["2022"]["cost_per_citation"])
    cpc_2024 = int(ym["2024"]["cost_per_citation"])
    perfte_2018 = round(float(ym["2018"]["citations_per_fte"]))
    perfte_2024 = round(float(ym["2024"]["citations_per_fte"]))
    cum_rev = sum(int(ym[y]["fee_revenue"]) for y in ym)
    cum_cost = sum(int(ym[y]["annual_cost"]) for y in ym)
    recovery = round(100 * cum_rev / cum_cost)
    base_cost = int(ym["2018"]["annual_cost"])  # 292,399 full-team baseline

    # first-offense range across all years
    fo = [float(ym[y]["first_offense_pct"]) for y in ym if ym[y]["first_offense_pct"]]
    fo_lo, fo_hi = round(min(fo)), round(max(fo))

    # per-park (named) counts
    pc = Counter(r["location_canon"] for r in named)
    top_parks = pc.most_common(5)  # Discovery 564, Magnuson 367, ...
    named_share = round(100 * len(named) / dlp_total)
    lincoln = pc.get("Lincoln Park", 0)

    # top-10 share, pre/post COVID, over named citations (matches the prose denominator)
    def top10_share(lo: int, hi: int) -> int:
        seg = [r for r in named if lo <= yr(r) <= hi]
        c = Counter(r["location_canon"] for r in seg)
        return round(100 * sum(n for _, n in c.most_common(10)) / len(seg))

    pre_share, post_share = top10_share(2014, 2019), top10_share(2020, 2026)

    # 2026 annualized + projection
    y2026 = dlp_by_year["2026"]
    factor = 365 / 107  # day-of-year cutoff 2026-04-17
    annualized_2026 = round(y2026 * factor)

    # OLA acreage (Finding 05)
    olas = {}
    if OLAS_CSV.exists():
        with OLAS_CSV.open() as fh:
            for o in csv.DictReader(fh):
                olas[o["ola_name"]] = o.get("acres", "")

    # ---- assertions: (label, substring that must appear verbatim) ----
    checks: list[tuple[str, str]] = [
        ("total DLP citations", f"{dlp_total:,}"),
        ("2018 peak", f"{peak_2018:,}"),
        ("best post-COVID year (2024)", f"{best_post}"),
        ("2024 as % of 2018 peak", f"{pct_2024_of_peak}%"),
        ("2014 baseline", "183"),
        ("2020 crater count", "393"),
        ("2020 drop vs 2019", f"{abs(drop_2020)}%"),
        ("cost/citation 2018 (low)", f"${cpc_2018:,}"),
        ("cost/citation 2022 (high)", f"${cpc_2022:,}"),
        ("cost/citation 2024", f"${cpc_2024}"),
        ("per-FTE 2018 peak", f"{perfte_2018}"),
        ("per-FTE 2024", f"{perfte_2024}"),
        ("cumulative fee revenue", f"${cum_rev:,}"),
        ("cost recovery %", f"{recovery}%"),
        ("cumulative cost $3.34M", "$3.34M"),
        ("first-offense range low", f"{fo_lo}"),
        ("first-offense range high", f"{fo_hi}"),
        ("non-DLP post-2019 rows", f"{non_dlp_post2019}"),
        ("2019 DLP (full year)", f"{dlp_by_year['2019']:,}"),
        ("named-park share", f"{named_share}%"),
        ("Lincoln Park cumulative", f"{lincoln}"),
        ("top-10 pre-COVID share", f"{pre_share}%"),
        ("top-10 post-COVID share", f"{post_share}%"),
        ("2026 annualized estimate", f"{annualized_2026}"),
        ("baseline cost $292,399", f"${base_cost:,}"),
        ("single-officer ACO cost $152,399", f"${int(ym['2018']['traceable_aco_cost']):,}"),
        ("three-officer ACO cost $454,652", f"${int(ym['2023']['funded_aco_cost']):,}"),
    ]
    # top-5 named parks with their counts (footnote: "Discovery 564, ...")
    park_short = {"Discovery Park": "Discovery", "Magnuson Park": "Magnuson",
                  "Volunteer Park": "Volunteer", "Woodland Park": "Woodland",
                  "Golden Gardens Park": "Golden Gardens"}
    for park, n in top_parks:
        label = park_short.get(park, park)
        checks.append((f"top park {label} count", f"{label} {n}"))
    # OLA acreages cited in Finding 05
    for name in ("Kinnear", "Magnolia Manor"):
        if name in olas:
            checks.append((f"{name} acreage", f"{olas[name]} acre"))

    for label, needle in checks:
        check(needle in html, f"prose: {label} ({needle!r}) present in site HTML")


# Map tile hosts that are dead or deprecated and must not appear in any public
# page. CARTO retired the fastly host (it now serves grey/blank tiles); the live
# endpoint is {s}.basemaps.cartocdn.com. Add hosts here as vendors rotate them.
DEAD_TILE_HOSTS = [
    "cartodb-basemaps-",          # old CARTO fastly host (light_all etc.)
    ".global.ssl.fastly.net",     # the bare fastly domain it lived on
]


def verify_asset_urls() -> None:
    """Fail if any public page references a known-dead map tile host."""
    print("\n[10] site HTML assets — no dead/deprecated map tile hosts")
    html, pages, _ = _site_corpus()
    if not html:
        print("  INFO  no public/staged HTML pages found — skipping asset check")
        return
    for name in PUBLIC_PAGES + STAGED_PAGES:
        p = DOCS_DIR / name
        if not p.exists():
            continue
        text = p.read_text()
        for host in DEAD_TILE_HOSTS:
            check(host not in text, f"asset: {name} free of dead tile host {host!r}")


# ---- Entry point ----------------------------------------------------------

def main() -> int:
    print("Enforcement data verification harness")
    print(f"Repo root: {REPO_ROOT}")

    verify_new_prr_raw()
    verify_old_prr_raw()

    rows = load_citations()
    verify_citations_csv(rows)
    verify_fee_arithmetic(rows)
    verify_hotspots(rows)
    verify_hotspots_extra(rows)
    verify_by_park_year(rows)
    verify_year_metrics(rows)
    verify_html_prose(rows)
    verify_asset_urls()

    print("\n" + "=" * 72)
    if failures:
        print(f"FAILURES: {len(failures)}")
        for f in failures:
            print(f"  - {f}")
        return 1
    print("ALL CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
