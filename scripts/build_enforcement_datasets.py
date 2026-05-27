#!/usr/bin/env python3
"""
Build the enforcement CSVs from the two SPR public-records-request responses.

Inputs:
  - data/prr-responses/C049204/*.xlsx
      Five workbooks produced by SPR on 2019-10-15 in response to PRR
      C049204. Format: one sheet per offense level per year-range, 14
      columns with a row-0 header. Covers 2014-01-01 → 2019-10-15.
      Scope: Parks Code "Dog Loose in Park" (DLP) violations only.

  - data/prr-responses/C263949/*.xlsx
      Four workbooks produced by Seattle FAS on 2026-05-08–05-15 in
      response to PRR C263949. Format: SSRS report export, one sheet
      per file, parameter rows 0-1, header row 3, 49 columns. Covers
      2019-01-01 → 2026-04-17. Scope: all parks-related violations
      (DLP + license + scoop + permit + vaccinate + other).

Outputs:
  - data/enforcement-citations.csv
      One row per violation across both PRRs, with deduplication
      handled at the year level (see below). Columns are
      backwards-compatible with the original v1 schema; new fields are
      appended at the end.

  - data/enforcement-by-park-year.csv
      Aggregate count by (location_canon, year, dlp_only). The
      `dlp_only` column allows downstream charts to filter to the
      apples-to-apples 2014–2026 DLP-only series, or to the broader
      all-violations 2019–2026 series, without re-reading the wide CSV.

2019 overlap rule
  C049204 and C263949 both contain 2019 records. C049204's 2019 sheet
  ends 2019-10-15; C263949 covers the full year. The two PRRs return
  different violation scopes (DLP-only vs. all-parks). Rather than
  attempt row-level dedup across two non-overlapping schemas, this
  script DROPS ALL 2019 ROWS FROM C049204 and uses C263949's full-year
  2019 as the authoritative source. The change is logged in the
  script output and verified by `scripts/verify_enforcement_data.py`.

Usage:
  cd <repo root> && .venv/bin/python scripts/build_enforcement_datasets.py

Dependencies: openpyxl. Install in a venv:
    python3 -m venv .venv && .venv/bin/pip install openpyxl
"""

from __future__ import annotations

import csv
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
PRR_OLD_DIR = REPO_ROOT / "data" / "prr-responses" / "C049204"
PRR_NEW_DIR = REPO_ROOT / "data" / "prr-responses" / "C263949"
CITATIONS_OUT = REPO_ROOT / "data" / "enforcement-citations.csv"
BY_PARK_YEAR_OUT = REPO_ROOT / "data" / "enforcement-by-park-year.csv"

# Field schema (column order in enforcement-citations.csv).
# IMPORTANT: do not reorder, only append. Several downstream tools and
# pre-existing chart consumers index by these names.
CSV_FIELDS = [
    "year",
    "offense_level",
    "location_raw",
    "location_canon",
    "zip",
    "issued_at",
    "fee",
    "case_result",
    "source_file",
    "source_sheet",
    "location_type",
    # New fields appended for the C263949 ingest (2026):
    "violation_item",
    "violation_category",
    "dlp_only",
    "district",
    "officer",
    "result_raw",
    "source_prr",
]

# Canonicalization map: regex → canonical name.
# Applied left-to-right; first match wins. Patterns are case-insensitive,
# anchored at the start of the address string (everything after the first
# comma or hyphen is typically a sub-location like "baseball diamond").
CANONICAL_PARKS = [
    (r"^Warren G\.?\s*Magnuson.*", "Magnuson Park"),
    (r"^Magnuson.*", "Magnuson Park"),
    (r"^Golden Gardens.*", "Golden Gardens Park"),
    (r"^Cal Anderson.*", "Cal Anderson Park"),
    (r"^Genesee Park.*", "Genesee Park"),
    (r"^Woodland Park.*", "Woodland Park"),
    (r"^Lower Woodland.*", "Woodland Park"),
    (r"^Green Lake.*", "Green Lake Park"),
    (r"^Volunteer Park.*", "Volunteer Park"),
    (r"^Discovery Park.*", "Discovery Park"),
    (r"^Lincoln Park.*", "Lincoln Park"),
    (r"^Seward Park.*", "Seward Park"),
    (r"^Martha Washington.*", "Martha Washington Park"),
    (r"^Westcrest.*", "Westcrest Park"),
    (r"^Ravenna Park.*", "Ravenna Park"),
    (r"^Wallingford Playfield.*", "Wallingford Playfield"),
    (r"^Rogers Playground.*", "Rogers Playground"),
    (r"^Maple Leaf Reservoir.*", "Maple Leaf Reservoir Park"),
    (r"^Meridian Playground.*", "Meridian Playground"),
    (r"^Soundview Playfield.*", "Soundview Playfield"),
    (r"^Ballard Playground.*", "Ballard Playground"),
    (r"^W\.?\s*Queen Anne Playfield.*", "West Queen Anne Playfield"),
    (r"^Kinnear.*", "Kinnear Park"),
    (r"^Denny Park.*", "Denny Park"),
    (r"^Gas Works.*", "Gas Works Park"),
    (r"^Matthews Beach.*", "Matthews Beach Park"),
    (r"^Madrona.*", "Madrona Park"),
    (r"^Kerry.*", "Kerry Park"),
    (r"^Alki Beach.*", "Alki Beach Park"),
    (r"^Dr\.?\s*Jose Rizal.*|^Jose Rizal.*", "Dr. Jose Rizal Park"),
    (r"^I-5 Colonnade.*|^I5 Colonnade.*", "I-5 Colonnade"),
    (r"^Blue Dog Pond.*", "Blue Dog Pond"),
    (r"^Northacres.*", "Northacres Park"),
    (r"^Gilman Play(?:ground|field).*", "Gilman Playground"),
    (r"^Carkeek.*", "Carkeek Park"),
    (r"^Laurelhurst Play(?:ground|field).*", "Laurelhurst Playfield"),
    (r"^David Rodgers.*|^David Rogers.*", "David Rodgers Park"),
    (r"^Salmon Bay.*", "Salmon Bay Park"),
    (r"^Myrtle Edwards.*", "Myrtle Edwards Park"),
    (r"^Smith Cove.*", "Smith Cove Park"),
    (r"^Interbay Play(?:ground|field).*", "Interbay Playfield"),
    (r"^Greenlake.*", "Green Lake Park"),
    (r"^Ross Play(?:ground|field).*", "Ross Playground"),
    (r"^Alki Play(?:ground|field).*", "Alki Playground"),
    (r"^Leschi.*", "Leschi Park"),
    (r"^Schmitz.*", "Schmitz Preserve Park"),
    (r"^Madison Park.*", "Madison Park"),
    (r"^Pratt.*", "Pratt Park"),
    (r"^Hiawatha.*", "Hiawatha Playfield"),
    (r"^Fairmount.*", "Fairmount Park"),
    (r"^Meridian Play(?:ground|field).*", "Meridian Playground"),
    (r"^Sandel.*", "Sandel Playground"),
    (r"^Montlake Play(?:ground|field).*", "Montlake Playfield"),
    (r"^Cascade Play(?:ground|field).*", "Cascade Playground"),
    (r"^Loyal Heights.*", "Loyal Heights Playfield"),
    (r"^Meadowbrook.*", "Meadowbrook Playfield"),
    (r"^Maple Leaf.*", "Maple Leaf Reservoir Park"),
    (r"^TT Minor.*|^T\.?T\.? Minor.*", "TT Minor Playground"),
    (r"^Ballard Comm(?:unity)? Center.*", "Ballard Community Center"),
    (r"^Lawton.*", "Lawton Park"),
    (r"^Wallingford Play(?:ground|field).*", "Wallingford Playfield"),
    (r"^Cowen.*", "Cowen Park"),
    (r"^Powell Barnett.*", "Powell Barnett Park"),
    (r"^Othello.*", "Othello Playground"),
    (r"^Beacon Hill Play(?:ground|field).*", "Beacon Hill Playfield"),
    (r"^Jefferson Park.*", "Jefferson Park"),
    (r"^Bhy Kracke.*", "Bhy Kracke Park"),
    (r"^Ella Bailey.*", "Ella Bailey Park"),
    (r"^Cormorant Cove.*", "Cormorant Cove Park"),
    (r"^Hamilton Viewpoint.*", "Hamilton Viewpoint Park"),
    (r"^Lake Union.*", "Lake Union Park"),
    (r"^Peppi's.*|^Peppis.*", "Peppi's Playground"),
    (r"^E\.?C\.? Hughes.*", "E.C. Hughes Playground"),
    (r"^Bitter Lake.*", "Bitter Lake Playground"),
    (r"^Licton Springs.*", "Licton Springs Park"),
    (r"^Mount Baker.*|^Mt\.? Baker.*", "Mount Baker Park"),
    (r"^Seacrest.*", "Seacrest Park"),
    (r"^Judkins.*", "Judkins Park"),
    (r"^Rainier Playfield.*", "Rainier Playfield"),
    (r"^Yesler.*", "Yesler Terrace Park"),
]


# C263949 (SSRS export) column indices. Derived from data inspection;
# kept here as a single source of truth.
NEW_COL_CASEID = 2
NEW_COL_VIOL_ITEM = 10
NEW_COL_FEE = 18
NEW_COL_OFFICER = 19
NEW_COL_ISSUE_DATE = 20
NEW_COL_RESULT = 21
NEW_COL_INCIDENT_DATE = 29
NEW_COL_DISTRICT = 34
NEW_COL_LOCATION = 42
NEW_COL_ZIP = 45
NEW_COL_SENTINEL = 48

# Categorize a violation_item text into one of a small number of buckets.
# Used to slice the consolidated CSV without string-matching at chart time.
VIOL_CATEGORY_PATTERNS = [
    (re.compile(r"Dog Loose in Park", re.IGNORECASE), "dog_loose_in_park"),
    (re.compile(r"Permit Animal at Large", re.IGNORECASE), "permit_at_large"),
    (re.compile(r"Vaccinate", re.IGNORECASE), "vaccinate"),
    (re.compile(r"Scoop", re.IGNORECASE), "scoop"),
    (re.compile(r"License", re.IGNORECASE), "license"),
    (re.compile(r"False Statement", re.IGNORECASE), "false_statement"),
    (re.compile(r"Voided Citation", re.IGNORECASE), "voided"),
]

OFFENSE_LEVEL_RE = re.compile(r"\((\d+)(?:st|nd|rd|th) Offense\)", re.IGNORECASE)


def canonicalize(raw: str | None) -> str:
    if raw is None:
        return ""
    s = re.sub(r"\s+", " ", str(raw).strip()).rstrip(".").rstrip(",")
    for pattern, canon in CANONICAL_PARKS:
        if re.match(pattern, s, re.IGNORECASE):
            return canon
    return s


def classify_location(raw: str, canon: str) -> str:
    raw = (raw or "").strip()
    canon = (canon or "").strip()
    if not raw and not canon:
        return "unknown"
    if canon and re.match(r"^\s*\d", canon):
        return "street_address"
    if canon:
        return "park_named"
    return "unknown"


def categorize_violation(violation_item: str) -> str:
    if not violation_item:
        return "unknown"
    for rx, label in VIOL_CATEGORY_PATTERNS:
        if rx.search(violation_item):
            return label
    return "other"


def offense_level_from_item(violation_item: str) -> int:
    if not violation_item:
        return 0
    m = OFFENSE_LEVEL_RE.search(violation_item)
    return int(m.group(1)) if m else 0


def _isodate(dt) -> str:
    if isinstance(dt, datetime):
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    return str(dt) if dt else ""


def _year_of(dt) -> str:
    if isinstance(dt, datetime):
        return str(dt.year)
    s = str(dt) if dt else ""
    return s[:4] if s and s[:4].isdigit() else ""


def normalize_case_result(result_raw: str | None) -> str:
    """Map old + new vocabulary into a small, stable set of values."""
    if not result_raw:
        return ""
    s = str(result_raw).strip()
    sl = s.lower()
    if sl in ("verbal", "warning"):
        return "Verbal"
    if sl in ("voided citation", "canceled citation"):
        return "Voided"
    if sl in ("dismissed", "stricken"):
        return "Dismissed"
    return "Citation"  # citation, pending, guilty — all are "officer wrote it"


# ---- C049204 loader (old PRR) ---------------------------------------------

OLD_SHEET_RE = re.compile(r"(\d)(?:st|nd|rd|th) Offense\s+(.+)")


def load_old_prr() -> list[dict]:
    rows: list[dict] = []
    for xlsx_path in sorted(PRR_OLD_DIR.glob("*.xlsx")):
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            m = OLD_SHEET_RE.match(sheet_name)
            if not m:
                continue
            offense_level = int(m.group(1))
            ws = wb[sheet_name]
            headers: list[str] | None = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c).strip() if c is not None else "" for c in row]
                    continue
                if all(c is None for c in row):
                    continue
                d = dict(zip(headers, row))
                addr_raw = str(d.get("Address") or "").strip()
                dt = d.get("Issue Date/Time")
                violation_item = str(d.get("Violation Item") or "").strip()
                # Old PRR is DLP-only by construction; if Violation Item is
                # blank (very rare), synthesize one from the sheet name to
                # keep downstream filters honest.
                if not violation_item:
                    suffix = {1: "1st", 2: "2nd", 3: "3rd", 4: "4th"}.get(offense_level, f"{offense_level}th")
                    violation_item = f"Parks Code - Dog Loose in Park ({suffix} Offense)"
                canon = canonicalize(addr_raw)
                category = categorize_violation(violation_item)
                result_raw = str(d.get("Case Result") or "").strip()
                rows.append(
                    {
                        "year": _year_of(dt),
                        "offense_level": offense_level,
                        "location_raw": addr_raw,
                        "location_canon": canon,
                        "zip": str(d.get("Zip Code") or "").strip(),
                        "issued_at": _isodate(dt),
                        "fee": d.get("Fee") if d.get("Fee") not in (None, "") else "",
                        "case_result": normalize_case_result(result_raw),
                        "source_file": xlsx_path.name,
                        "source_sheet": sheet_name,
                        "location_type": classify_location(addr_raw, canon),
                        "violation_item": violation_item,
                        "violation_category": category,
                        "dlp_only": "True" if category == "dog_loose_in_park" else "False",
                        "district": "",
                        "officer": "",
                        "result_raw": result_raw,
                        "source_prr": "C049204",
                    }
                )
        wb.close()
    return rows


# ---- C263949 loader (new PRR, SSRS export) --------------------------------

def load_new_prr() -> tuple[list[dict], dict[str, int]]:
    """Returns (rows, per_file_sentinel) so the caller can verify totals."""
    rows: list[dict] = []
    sentinels: dict[str, int] = {}
    for xlsx_path in sorted(PRR_NEW_DIR.glob("*.xlsx")):
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        file_sentinel: int | None = None
        file_count = 0
        for raw_row in ws.iter_rows(values_only=True):
            if not raw_row or len(raw_row) < NEW_COL_SENTINEL + 1:
                continue
            if not isinstance(raw_row[NEW_COL_CASEID], (int, float)):
                continue
            # Record the sentinel value (same on every row of a given file).
            sentinel_cell = raw_row[NEW_COL_SENTINEL]
            if isinstance(sentinel_cell, str) and sentinel_cell.startswith("Total Violations:"):
                try:
                    file_sentinel = int(sentinel_cell.split(":")[1].strip())
                except (IndexError, ValueError):
                    pass

            issue_dt = raw_row[NEW_COL_ISSUE_DATE]
            incident_dt = raw_row[NEW_COL_INCIDENT_DATE]
            year_dt = issue_dt if isinstance(issue_dt, datetime) else incident_dt
            location_raw = str(raw_row[NEW_COL_LOCATION] or "").strip()
            violation_item = str(raw_row[NEW_COL_VIOL_ITEM] or "").strip()
            canon = canonicalize(location_raw)
            category = categorize_violation(violation_item)
            offense_level = offense_level_from_item(violation_item)
            zip_cell = raw_row[NEW_COL_ZIP]
            zip_str = str(int(zip_cell)) if isinstance(zip_cell, (int, float)) else str(zip_cell or "").strip()
            result_raw = str(raw_row[NEW_COL_RESULT] or "").strip()
            fee = raw_row[NEW_COL_FEE]
            if fee in (None, ""):
                fee = ""

            rows.append(
                {
                    "year": _year_of(year_dt),
                    "offense_level": offense_level if offense_level else "",
                    "location_raw": location_raw,
                    "location_canon": canon,
                    "zip": zip_str,
                    "issued_at": _isodate(issue_dt),
                    "fee": fee,
                    "case_result": normalize_case_result(result_raw),
                    "source_file": xlsx_path.name,
                    "source_sheet": ws.title,
                    "location_type": classify_location(location_raw, canon),
                    "violation_item": violation_item,
                    "violation_category": category,
                    "dlp_only": "True" if category == "dog_loose_in_park" else "False",
                    "district": str(raw_row[NEW_COL_DISTRICT] or "").strip(),
                    "officer": str(raw_row[NEW_COL_OFFICER] or "").strip(),
                    "result_raw": result_raw,
                    "source_prr": "C263949",
                }
            )
            file_count += 1
        wb.close()
        if file_sentinel is None:
            raise RuntimeError(f"No Total Violations sentinel found in {xlsx_path.name}")
        if file_sentinel != file_count:
            raise RuntimeError(
                f"{xlsx_path.name}: sentinel says {file_sentinel}, ingested {file_count} rows"
            )
        sentinels[xlsx_path.name] = file_sentinel
    return rows, sentinels


# ---- Main pipeline --------------------------------------------------------

def main() -> None:
    old_rows = load_old_prr()
    new_rows, new_sentinels = load_new_prr()

    # Apply 2019 overlap rule: drop old PRR's 2019 rows; new PRR's full-year
    # 2019 supersedes them.
    old_2019 = [r for r in old_rows if r["year"] == "2019"]
    old_kept = [r for r in old_rows if r["year"] != "2019"]
    print(
        f"2019 overlap: dropped {len(old_2019)} rows from C049204 (Jan 1 – Oct 15 2019), "
        f"using C263949 full-year 2019 as authoritative."
    )

    rows = old_kept + new_rows

    # Sort for stable diffs across re-runs.
    rows.sort(key=lambda r: (r["year"] or "", r["source_prr"], r["source_file"], r["source_sheet"], r["issued_at"]))

    # Write the consolidated CSV.
    CITATIONS_OUT.parent.mkdir(parents=True, exist_ok=True)
    with CITATIONS_OUT.open("w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=CSV_FIELDS)
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, "") for k in CSV_FIELDS})
    print(f"Wrote {CITATIONS_OUT.relative_to(REPO_ROOT)} ({len(rows)} rows).")

    # Aggregate by (location_canon, year, dlp_only).
    park_year_counts: Counter = Counter()
    for r in rows:
        if r["location_canon"] and r["year"]:
            park_year_counts[(r["location_canon"], r["year"], r["dlp_only"])] += 1
    with BY_PARK_YEAR_OUT.open("w", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["park", "year", "dlp_only", "citations"])
        for (park, year, dlp), n in sorted(park_year_counts.items()):
            writer.writerow([park, year, dlp, n])
    print(f"Wrote {BY_PARK_YEAR_OUT.relative_to(REPO_ROOT)}.")

    # Verification summary (matches checks done by verify_enforcement_data.py).
    print()
    print("=" * 72)
    print("VERIFICATION SUMMARY")
    print("=" * 72)
    print(f"Total rows: {len(rows)}  (old kept: {len(old_kept)}, new: {len(new_rows)})")

    print("\nPer-file C263949 sentinels:")
    for name, n in new_sentinels.items():
        print(f"  {name}: {n}")

    yc = Counter(r["year"] for r in rows if r["year"])
    print("\nBy year (all categories):")
    for y in sorted(yc):
        print(f"  {y}: {yc[y]}")

    yc_dlp = Counter(r["year"] for r in rows if r["year"] and r["dlp_only"] == "True")
    print("\nBy year (DLP only):")
    for y in sorted(yc_dlp):
        print(f"  {y}: {yc_dlp[y]}")

    cat_c = Counter(r["violation_category"] for r in rows)
    print("\nBy violation category:")
    for k, v in cat_c.most_common():
        print(f"  {k}: {v}")

    loc_c = Counter(r["location_canon"] for r in rows if r["location_canon"])
    print(f"\nUnique canonical locations: {len(loc_c)}")
    print("Top 10 locations (all years, all categories):")
    for loc, n in loc_c.most_common(10):
        print(f"  {n:5d}  {loc}")


if __name__ == "__main__":
    main()
